import openpyxl

wb = openpyxl.load_workbook('..\素材\Excel\produceSales.xlsx')
Garlic = 3.07
Celery = 1.19
Lemon = 1.27

sheet = wb.active
for row in sheet.rows:
    if row[0].value == 'Garlic':
        sheet.cell(row=row[0].row, column=2,value=Garlic)
    elif row[0].value == 'Celery':
        sheet.cell(row=row[0].row, column=2,value=Celery)
    elif row[0].value == 'Lemon':
        sheet.cell(row=row[0].row, column=2,value=Lemon)

wb.save('..\素材\Excel\\newProduceSales.xlsx')
