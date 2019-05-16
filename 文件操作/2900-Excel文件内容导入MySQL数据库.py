import os
import random
import string

import openpyxl

import pymysql
import configparser

for x in range(50):
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(1, random.randint(5,10)):
        for j in range(1, 6):
            value = ''.join(random.sample(string.ascii_letters + string.digits, 30))
            ws.cell(row=i, column=j, value=value)
    wb.save('../素材/Excel/Temp/t' + str(x) + '.xlsx')



config = configparser.ConfigParser()
config.read('configure.ini')
host = config['db']['host']
port = config.getint('db', 'port')
user = config['db']['user']
password = config['db']['password']
database = config['db']['database']
charset = config['db']['charset']

connect = pymysql.connect(host=host,
                          port=port,
                          user=user,
                          password=password,
                          database=database,
                          charset=charset)
cursor = connect.cursor()
filenames = os.listdir('../素材/Excel/Temp')
os.chdir('../素材/Excel/Temp')
for filename in filenames:
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    data = []
    for row in ws.rows:
        rowData = []
        for cell in row:
            rowData.append(cell.value)
        data.append(tuple(rowData))
    sql = "insert into ExcelToMysql value(%s,%s,%s,%s,%s)"
    result = cursor.executemany(sql, data)
    connect.commit()
    print(filename + '导入成功！')
cursor.close()
connect.close()

